import openpyxl
libro =openpyxl.load_workbook('frutería.xlsx')
hoja1=libro['ventas']




frutas={}
for i in range(2,8,1):
    nombre=hoja1.cell(i,1)
    precio=hoja1.cell(i,2)
    frutas[nombre.value]=precio.value
total=0
print(frutas)
opcion='s'
while opcion.lower()=="s":
    fruta=input('Que fruta quieres comprar?')
    if fruta.lower() not in frutas:
        print('fruta invalida')
num=eval(input('Cuantas frutas quieres'))
for i in range(2,8,1):
    nombre=hoja1.cell[i,1]
    precio=hoja1.cell[i,2]
    precio=eval(input('Inrgese la precio de la fruta'))
    hoja1[f'A{i+2}']=nombre
    hoja1[f'B{i+2}']=precio


libro.save('frutería.xlsx')
