import openpyxl
from openpyxl import load_workbook
with open("direccion.txt") as archivo:
    for linea in archivo:
        if len(linea)>1:
                print(linea)
        else:
            print('no tiene info')

directionTxt=linea
bookOneDrive=str(directionTxt)
characters = "="
for x in range(len(characters)):
    bookOneDrive = bookOneDrive.replace(characters[x],"")
print(bookOneDrive)



FILE_PATH=bookOneDrive
SHEET='DB Excel'
workbook=load_workbook(FILE_PATH,read_only=True)
sheet=workbook[SHEET]
Questions=[]
answer=[]
for row in sheet.iter_rows():
    value=row[3].value
    Questions.append(value)
print(Questions)
for i in range(len(Questions)):
    for j in range(len(Questions[i])):
        Q=(Questions[i])[j]
        if Q=="\n":
            answer.append(Q)
print(answer)
